"""
Canon imageRUNNER C3922i - Print Log Collector + Syslog Receiver
Ishlatish: python canon_server.py
Brauzerda: http://localhost:5000

Syslog sozlash:
  Canon Remote UI → Export/Clear Audit Log → Syslog Settings
  Use Syslog Send: ✅
  Syslog Server Address: <bu PC ning IP manzili>
  Syslog Server Port Number: 5140
  Connection Type: UDP
"""

import os, re, csv, io, json, socket, sqlite3, threading, logging, random, time
import requests
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, send_file, make_response
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("canon")

app = Flask(__name__)
CORS(app)

DB_PATH = "canon_logs.db"
SYSLOG_PORT = 5140          # Printerda ham shu port kiritiladi
SYSLOG_HOST = "0.0.0.0"

# Printers configuration is stored in the sqlite 'printers' table.

# ─── DATABASE ────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS print_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            printer_ip  TEXT,
            printer_name TEXT,
            job_no      TEXT,
            result      TEXT DEFAULT 'OK',
            start_time  TEXT,
            end_time    TEXT,
            job_type    TEXT DEFAULT 'Printer Print',
            file_name   TEXT,
            user_name   TEXT,
            original_pages INTEGER DEFAULT 0,
            output_pages   INTEGER DEFAULT 0,
            sheets      INTEGER DEFAULT 0,
            copies      INTEGER DEFAULT 1,
            source      TEXT DEFAULT 'csv',
            hidden      INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    # Migratsiya: agar hidden ustuni bo'lmasa, qo'shamiz
    try:
        c.execute("ALTER TABLE print_logs ADD COLUMN hidden INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    # Syslog live feed uchun alohida jadval
    c.execute("""
        CREATE TABLE IF NOT EXISTS syslog_raw (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            received  TEXT DEFAULT (datetime('now','localtime')),
            sender_ip TEXT,
            raw       TEXT
        )
    """)
    # Printers jadvali (CRUD uchun)
    c.execute("""
        CREATE TABLE IF NOT EXISTS printers (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            ip         TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    c.execute("SELECT COUNT(*) FROM printers")
    if c.fetchone()[0] == 0:
        default_printers = [
            ("Canon C3922i #1", "172.16.0.7"),
            ("Canon C3922i #2", "172.16.0.8"),
            ("Canon C3922i #3", "172.16.0.9"),
            ("Canon C3922i #4", "172.16.0.24"),
            ("Canon C3922i #5", "172.18.0.120"),
        ]
        c.executemany("INSERT INTO printers (name, ip) VALUES (?,?)", default_printers)
    conn.commit()
    conn.close()

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ─── UTILITIES ───────────────────────────────────────────────────────

# Canon printer bu nomlarni generic yuboradi — haqiqiy user emas
_GENERIC_USERS = {"user", "admin", "administrator", "guest", "unknown", "--------", ""}

USERNAME_MAP = {
    "o.atayeva": "Atayeva Ogulnur",
    "a.axadov": "Axadov Asror",
    "a.axmedov": "Axmedov Azizbek",
    "x.axundjanov": "Axundjanov Xumoyunmirzo",
    "u.azamatov": "Azamatov Uchqun",
    "i.babaxanov": "Babaxanov Ilhom",
    "sh.baxriddinov": "Baxriddinov Shamshodbek",
    "sh.baxtiyarov": "Baxtiyarov Shaxzod",
    "b.muhammadsaidov": "Baxtiyor Muxammadsaidov",
    "sh.beknazarov": "Beknazarov Sherzod",
    "s.bobokulov": "Bobokulov Sodiq",
    "d.bulanov": "Bulanov Dilshodbek",
    "m.butayev": "Butayev Ma'ruf",
    "a.abdullayev": "Abdullayev Abbas",
    "a.abduraximov": "Abduraximov Abdulaziz",
    "sh.abdurahmanov": "Abduraxmanov Shuxrat",
    "sh.asatullayeva": "Asatullayeva Shirin",
    "s.solijonov": "Solijonov Sarvar",
    "b.meliyev": "Meliyev Baxtiyor",
    "f.ergashev": "Ergashev Farrux",
    "m.ergashev": "Ergashev Mirzobek",
    "r.eshniyozov": "Eshniyozov Rustambek",
    "i.gaporov": "Gaporov Isfandiyorbek",
    "o.ibroximov": "Ibroximov Otabek",
    "k.isakova": "Isakova Kamila",
    "b.isoyev": "Isoyev Baxodir",
    "i.isomiddinov": "Isomiddinov Ibodullo",
    "f.istamov": "Istamov Firdavs",
    "r.jumaboyev": "Jumaboyev Rashid",
    "m.karimov": "Karimov Ma'murbek",
    "u.karimov": "Karimov Ulug'bek",
    "m.kodirov": "Kodirov Maruf",
    "b.madraximov": "Madraximov Baxtiyorjon",
    "a.mavlonov": "Mavlonov Akmal",
    "o.mirzayev": "Mirzayev Otabek",
    "b.musayev": "Musayev Boburjon",
    "a.muxammatov": "Muxammatov Alisher",
    "o.nabiev": "Nabiev Omonulla",
}

def parse_username(raw):
    """DOMAIN\\user.name yoki user.name → 'User Name'"""
    if not raw:
        return "Noma'lum"
    raw = raw.strip()
    if "\\" in raw:
        raw = raw.split("\\")[-1]
    raw = re.sub(r"@.*$", "", raw)

    key = raw.lower()
    if key in USERNAME_MAP:
        return USERNAME_MAP[key]

    # Generic nomlarni qaytarish (Department ID autentifikatsiya)
    if raw.lower() in _GENERIC_USERS:
        return raw.strip() or "Noma'lum"

    parts = re.split(r"[._\-\s]+", raw)
    return " ".join(p.capitalize() for p in parts if p) or raw

def printer_by_ip(ip):
    conn = get_conn()
    row = conn.execute("SELECT name, ip FROM printers WHERE ip = ?", (ip,)).fetchone()
    conn.close()
    if row:
        return {"name": row["name"], "ip": row["ip"]}
    return {"name": f"Printer ({ip})", "ip": ip}

def parse_audit_csv(content, printer):
    """Canon Audit Log CSV formatini parse qilish.
    Audit log da haqiqiy username bo'ladi (b.meliyev kabi).
    Ustunlar: Date/Time, Result, Function, User Name, ...
    """
    logs = []
    try:
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            return logs
        # Audit log ekanligini tekshirish
        fields_lower = [f.lower() for f in reader.fieldnames if f]
        if "function" not in " ".join(fields_lower) and "audit" not in " ".join(fields_lower):
            return logs  # Bu audit log emas

        for row in reader:
            try:
                def pick(row, *keys):
                    for k in keys:
                        for field, val in row.items():
                            if field and k.lower() in field.lower():
                                return (val or "").strip()
                    return ""

                func = pick(row, "function", "operation")
                if "print" not in func.lower() and "copy" not in func.lower():
                    continue

                raw_user = pick(row, "user name", "username", "user")
                result   = pick(row, "result") or "OK"
                dt       = pick(row, "date", "time", "datetime")
                pages_s  = pick(row, "page", "sheet")
                pages    = int(pages_s) if pages_s.isdigit() else 1

                if not raw_user or raw_user.lower() in _GENERIC_USERS:
                    continue

                logs.append({
                    "printer_ip":    printer["ip"],
                    "printer_name":  printer["name"],
                    "job_no":        "",
                    "result":        result,
                    "start_time":    dt,
                    "end_time":      dt,
                    "job_type":      func or "Printer Print",
                    "file_name":     pick(row, "file", "document", "job name"),
                    "user_name":     parse_username(raw_user),
                    "original_pages": pages,
                    "output_pages":   pages,
                    "sheets":        pages,
                    "copies":        1,
                })
            except Exception:
                continue
    except Exception as e:
        log.warning(f"Audit CSV parse xatolik: {e}")
    return logs


def save_logs(logs, source="csv"):
    if not logs:
        return 0
    conn = get_conn()
    c = conn.cursor()
    saved = 0
    for lg in logs:
        # Job allaqachon mavjudligini tekshirish
        exists = c.execute(
            "SELECT id, source FROM print_logs WHERE printer_ip=? AND job_no=? AND start_time=?",
            (lg["printer_ip"], lg.get("job_no",""), lg.get("start_time",""))
        ).fetchone()
        if not exists:
            c.execute("""
                INSERT INTO print_logs
                (printer_ip,printer_name,job_no,result,start_time,end_time,
                 job_type,file_name,user_name,original_pages,output_pages,sheets,copies,source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                lg["printer_ip"], lg.get("printer_name",""), lg.get("job_no",""),
                lg.get("result","OK"), lg.get("start_time",""), lg.get("end_time",""),
                lg.get("job_type","Printer Print"), lg.get("file_name",""),
                lg.get("user_name","Noma'lum"),
                lg.get("original_pages",0), lg.get("output_pages",0),
                lg.get("sheets",0), lg.get("copies",1), source
            ))
            saved += 1
        elif source == "csv" and exists["source"] == "syslog":
            # Syslog yozuvini CSV dagi aniq sahifa va nusxa ma'lumotlari bilan yangilaymiz
            c.execute("""
                UPDATE print_logs
                SET original_pages=?, output_pages=?, sheets=?, copies=?, file_name=?, source='csv', hidden=0
                WHERE id=?
            """, (
                lg.get("original_pages", 0), lg.get("output_pages", 0),
                lg.get("sheets", 0), lg.get("copies", 1), lg.get("file_name", ""),
                exists["id"]
            ))
            saved += 1
    conn.commit()
    conn.close()
    return saved


# ─── CSV PARSER ──────────────────────────────────────────────────────

def parse_csv_content(content, printer):
    logs = []
    try:
        reader = csv.DictReader(io.StringIO(content))
    except Exception:
        return logs

    # Ustun nomlarini flexible moslashtirish
    def pick(row, *keys):
        for k in keys:
            for field, val in row.items():
                if field and k.lower() in field.lower():
                    return val
        return ""

    for row in reader:
        try:
            job_no    = pick(row, "job no", "job_no", "no.")
            result    = pick(row, "result") or "OK"
            start_t   = pick(row, "start time", "start_time")
            end_t     = pick(row, "end time", "end_time")
            job_type  = pick(row, "job type", "type") or "Printer Print"
            file_name = pick(row, "file name", "filename", "file")
            raw_user  = pick(row, "user name", "username", "user")
            orig      = int(pick(row, "original page") or 0)
            out_p     = int(pick(row, "output page") or 0)
            sheets_raw = pick(row, "sheet") or "1x1"

            if "x" in str(sheets_raw).lower():
                pts = str(sheets_raw).lower().split("x")
                sheets = int(pts[0]) if pts[0].strip().isdigit() else 1
                copies = int(pts[1]) if len(pts)>1 and pts[1].strip().isdigit() else 1
            else:
                sheets = int(sheets_raw) if str(sheets_raw).strip().isdigit() else 1
                copies = 1

            if not job_no and not file_name:
                continue

            logs.append({
                "printer_ip":    printer["ip"],
                "printer_name":  printer["name"],
                "job_no":        str(job_no),
                "result":        str(result).strip(),
                "start_time":    str(start_t),
                "end_time":      str(end_t),
                "job_type":      str(job_type),
                "file_name":     str(file_name),
                "user_name":     parse_username(raw_user),
                "original_pages": orig,
                "output_pages":   out_p,
                "sheets":        sheets,
                "copies":        copies,
            })
        except Exception:
            continue
    return logs

# ─── REMOTE CSV FETCH ────────────────────────────────────────────────

PRINTER_USER = "Administrator"
PRINTER_PASS = "12345678"


def fetch_printer_logs(printer):
    ip = printer["ip"]
    session = requests.Session()

    # 1. Root sahifani yuklab Challenge va PK ni qidiramiz (yangi login formati)
    challenge, pk_pem = None, None
    try:
        r_get = session.get(f"http://{ip}:8000/", timeout=10)
        if r_get.status_code == 200:
            challenge_match = re.search(r'id\s*=\s*"CHALLENGE"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
            if not challenge_match:
                challenge_match = re.search(r'name\s*=\s*"CHALLENGE"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
            challenge = challenge_match.group(1) if challenge_match else None

            pk_match = re.search(r'id\s*=\s*"PK"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
            if not pk_match:
                pk_match = re.search(r'name\s*=\s*"PK"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
            pk_pem = pk_match.group(1) if pk_match else None
    except Exception as e:
        log.warning(f"Root sahifani yuklashda xatolik {ip}: {e}")

    # 2. Login qilish
    if challenge and pk_pem:
        # Yangi format: RSA-PKCS1v1.5 shifrlangan login
        try:
            import base64
            from cryptography.hazmat.primitives.asymmetric import padding
            from cryptography.hazmat.primitives import serialization
            
            public_key = serialization.load_pem_public_key(pk_pem.strip().replace("\r", "").encode('utf-8'))
            message = (PRINTER_PASS + challenge).encode('utf-8')
            ciphertext = public_key.encrypt(message, padding.PKCS1v15())
            encrypted_pass = base64.b64encode(ciphertext).decode('utf-8')
            
            login_data = {
                "USERNAME": PRINTER_USER,
                "PASSWORD": encrypted_pass,
                "PASSWORD_T": "",
                "CHALLENGE": challenge,
                "URI": "/",
                "policy": "",
                "DOMAIN": "localhost",
                "admin": "",
                "GUEST": "",
                "invalidCH": ""
            }
            login_resp = session.post(f"http://{ip}:8000/login", data=login_data, timeout=10)
            log.info(f"RSA Login {ip}: HTTP {login_resp.status_code}")
        except Exception as e:
            log.warning(f"RSA Login urinishda xatolik {ip}: {e}")
    else:
        # Agar yangi format topilmasa, eski /login.cgi usulini sinab ko'ramiz
        log.info(f"Login challenge topilmadi {ip}, eski login.cgi urinib ko'rilmoqda...")
        try:
            login_resp = session.post(
                f"http://{ip}:8000/login.cgi",
                data={"UserName": PRINTER_USER, "Password": PRINTER_PASS, "Action": "Login"},
                timeout=10
            )
            log.info(f"Eski Login {ip}: HTTP {login_resp.status_code}")
        except Exception as e:
            log.warning(f"Eski Login urinishda xatolik {ip}: {e}")

    # 3. Print sub-tizim sessiyasini faollashtirish (iR cookieni olish)
    try:
        r_native = session.get(f"http://{ip}:8000/rps/nativetop.cgi", timeout=10)
        log.info(f"Nativetop {ip}: HTTP {r_native.status_code}")
    except Exception as e:
        log.warning(f"Nativetop yuklashda xatolik {ip}: {e}")

    # 4. CSV yuklash — bir nechta URL variantini sinab ko'rish
    # Canon printer konfiguratsiyasiga qarab turli URL ishlashi mumkin
    dummy = int(datetime.now().timestamp() * 1000)
    urls_to_try = [
        # Asosiy Copy/Print log (CorePGTAG=2)
        f"http://{ip}:8000/rps/jlp.cgi?Flag=Csv_Data&LogType=0&CorePGTAG=2&Dummy={dummy}",
        # Audit log (haqiqiy username bilan)
        f"http://{ip}:8000/auditlog/AuditLog.cgi?Flag=Csv_Data&Dummy={dummy}",
        # Send log (CorePGTAG=3)
        f"http://{ip}:8000/rps/jlp.cgi?Flag=Csv_Data&LogType=0&CorePGTAG=3&Dummy={dummy}",
    ]

    for url in urls_to_try:
        try:
            r = session.get(url, timeout=20)
            if r.status_code != 200:
                continue
            text = r.content.decode("utf-8-sig", errors="replace")

            # Agar login sahifasiga yoki xatolikka qaytargan bo'lsa o'tkazib yuboramiz
            if "login" in text.lower() and len(text) < 2000:
                continue
            if "cannot open this page" in text.lower() and len(text) < 3000:
                continue

            # Bo'sh yoki CSV formatida bo'lmasa o'tkazib yuborish
            if len(text) < 50 or ("," not in text and ";" not in text):
                continue

            log.info(f"CSV olindi {ip}: {len(text)} bayt, URL: {url.split('?')[1][:40]}")
            return text, None

        except requests.exceptions.ConnectionError:
            return None, f"{ip} — ulanib bo'lmadi"
        except requests.exceptions.Timeout:
            continue
        except Exception as e:
            log.warning(f"URL xatolik {ip}: {e}")
            continue

    return None, f"{ip} — CSV yuklab bo'lmadi (barcha URL sinab ko'rildi)"

# ─── SYSLOG RECEIVER ─────────────────────────────────────────────────
#
# Canon C3922i RFC-5424 / RFC-3164 syslog formatida yubora oladi.
# Misol paket:
#   <14>1 2026-06-01T10:23:45Z 172.16.0.7 Canon - - - Print Job: User=s.solijonov Pages=4 Sheets=2 File=report.docx
#   <134>Jun  1 10:23:45 172.16.0.7 print: user=admin pages=1
#
# Biz barcha kelgan paketni saqlaymiz (syslog_raw),
# va agar ichida "print" + "user" + "page" bo'lsa – print_logs ga ham qo'shamiz.

_SYSLOG_PATTERNS = [
    # Canon Audit Log format
    re.compile(
        r'(?:User|UserName)[=:\s]+"?([^\s",]+)"?.*?'
        r'(?:Pages?|OutputPages?)[=:\s]+(\d+)',
        re.I
    ),
    # Fallback: user= pages=
    re.compile(
        r'user[=:\s]+"?([^\s",]+)"?.*?pages?[=:\s]+(\d+)',
        re.I
    ),
]

_FILE_RE  = re.compile(r'(?:File|FileName|Document)[=:\s]+"?([^",\r\n]+)"?', re.I)
_JOB_RE   = re.compile(r'(?:JobNo|Job\.?No\.?|JobID)[=:\s]+(\d+)', re.I)
_SHEET_RE = re.compile(r'(?:Sheets?|SheetCount)[=:\s]+(\d+)', re.I)


# Syslog xabaridan haqiqiy printer IP sini ajratib olish
_IP_RE = re.compile(r'\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b')

def extract_printer_ip_from_syslog(data: str, sender_ip: str) -> str:
    """Syslog xabaridan haqiqiy printer IP sini ajratib oladi.

    RFC 5424: <PRI>VERSION TIMESTAMP HOSTNAME APPNAME ...
      -> <14>1 2026-06-01T10:23:45Z 172.16.0.7 Canon ...
    RFC 3164: <PRI>TIMESTAMP HOSTNAME MSG
      -> <134>Jun  1 10:23:45 172.16.0.7 print: ...

    Agar xabarda IP topilmasa, sender_ip qaytariladi.
    """
    try:
        # <...> priority/version qismini olib tashlaymiz
        text = re.sub(r'^<\d+>\d*\s*', '', data).strip()

        tokens = text.split()
        # RFC 5424: birinchi token — timestamp (T harfi yoki Z bilan tugaydi)
        # Ikkinchi token — hostname (printer IP)
        if tokens and ('T' in tokens[0] or tokens[0].endswith('Z')):
            # RFC 5424: TIMESTAMP HOSTNAME ...
            if len(tokens) >= 2:
                candidate = tokens[1]
                if _IP_RE.fullmatch(candidate):
                    return candidate
        else:
            # RFC 3164: "Mon DD HH:MM:SS HOSTNAME ..."
            # Odatda 3 ta sana tokeni (Mmm DD HH:MM:SS), keyin hostname
            if len(tokens) >= 4:
                candidate = tokens[3]
                if _IP_RE.fullmatch(candidate):
                    return candidate

        # Fallback: xabar ichidagi birinchi 172.x.x.x yoki 10.x.x.x IP
        for m in _IP_RE.finditer(data):
            ip = m.group(1)
            parts = ip.split('.')
            # Gateway/loopback emasligini tekshiramiz
            if ip == sender_ip:
                continue
            if ip.startswith('127.') or ip == '0.0.0.0':
                continue
            # RFC1918 private IP — printer bo'lishi ehtimoli yuqori
            if parts[0] in ('10', '172', '192'):
                return ip
    except Exception:
        pass
    return sender_ip


def parse_syslog_packet(data: str, sender_ip: str):
    """Syslog paketidan print ma'lumotlarini ajratib olish.
    Ham standart key-value formatini, ham Canon Audit Log (CSV) formatini qo'llab-quvvatlaydi.
    """
    lower = data.lower()

    # Haqiqiy printer IP sini xabardan aniqlaymiz (gateway orqali kelsa sender_ip noto'g'ri bo'ladi)
    real_ip = extract_printer_ip_from_syslog(data, sender_ip)
    if real_ip != sender_ip:
        log.info(f"📡 Syslog: real printer IP={real_ip} (sender={sender_ip} — gateway orqali keldi)")

    # 1. Canon Audit Log (CSV) formatini tekshirish
    r_idx = data.find(' - ')
    if r_idx != -1:
        csv_part = data[r_idx + 3:].lstrip('\ufeff')
        if ',' in csv_part:
            parts = [p.strip() for p in csv_part.split(',')]
            if len(parts) >= 16:
                log_code = parts[0]
                op_type = parts[9]

                # '1001' — chop etish, '1002' — nusxa olish audit log kodi
                if op_type.lower() == 'print' or op_type.lower() == 'copy' or log_code in ('1001', '1002'):
                    user_name = parts[5]
                    result = parts[7] or "OK"
                    job_no = parts[12]
                    start_time = parts[13]
                    end_time = parts[14]
                    file_name = parts[15]

                    if not start_time:
                        start_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    if not end_time:
                        end_time = start_time

                    printer = printer_by_ip(real_ip)

                    return {
                        "printer_ip":    real_ip,
                        "printer_name":  printer["name"],
                        "job_no":        job_no,
                        "result":        result,
                        "start_time":    start_time,
                        "end_time":      end_time,
                        "job_type":      "Printer Print" if op_type.lower() == 'print' else "Copy",
                        "file_name":     file_name,
                        "user_name":     parse_username(user_name),
                        "original_pages": 1,
                        "output_pages":   1,
                        "sheets":        1,
                        "copies":        1,
                    }

    # 2. Key-value format (Standart syslog) fallback
    if "print" not in lower and "copy" not in lower:
        return None

    user_name, pages = None, 0
    for pat in _SYSLOG_PATTERNS:
        m = pat.search(data)
        if m:
            user_name = m.group(1)
            pages = int(m.group(2))
            break

    if not user_name:
        return None

    file_m  = _FILE_RE.search(data)
    job_m   = _JOB_RE.search(data)
    sheet_m = _SHEET_RE.search(data)

    printer = printer_by_ip(real_ip)
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    return {
        "printer_ip":    real_ip,
        "printer_name":  printer["name"],
        "job_no":        job_m.group(1) if job_m else "",
        "result":        "OK",
        "start_time":    now,
        "end_time":      now,
        "job_type":      "Printer Print" if "print" in lower else "Copy",
        "file_name":     file_m.group(1).strip() if file_m else "",
        "user_name":     parse_username(user_name),
        "original_pages": pages,
        "output_pages":   pages,
        "sheets":        int(sheet_m.group(1)) if sheet_m else pages,
        "copies":        1,
    }



def syslog_server():
    """UDP Syslog tinglash daemon (alohida thread da ishlaydi)"""
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.bind((SYSLOG_HOST, SYSLOG_PORT))
        log.info(f"✅ Syslog server UDP:{SYSLOG_PORT} da ishga tushdi")
    except OSError as e:
        log.error(f"❌ Syslog port {SYSLOG_PORT} band yoki ruxsat yo'q: {e}")
        log.error("   Canon printerda boshqa port sozlang yoki administratorlik bilan ishga tushiring")
        return

    while True:
        try:
            data, addr = sock.recvfrom(4096)
            sender_ip = addr[0]
            text = data.decode("utf-8", errors="replace").strip()

            # Xom ma'lumotni saqlash
            conn = get_conn()
            conn.execute(
                "INSERT INTO syslog_raw (sender_ip, raw) VALUES (?,?)",
                (sender_ip, text[:2000])
            )
            conn.commit()
            conn.close()

            # Print log parse qilish
            parsed = parse_syslog_packet(text, sender_ip)
            if parsed:
                saved = save_logs([parsed], source="syslog")
                if saved:
                    log.info(f"📄 Syslog log saqlandi: {parsed['user_name']} — {parsed['sheets']} varaq ({sender_ip})")
        except Exception as e:
            log.error(f"Syslog xatolik: {e}")


# ─── FLASK API ───────────────────────────────────────────────────────

@app.route("/")
def index():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "canon_dashboard.html")
    with open(path, encoding="utf-8") as f:
        response = make_response(f.read())
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response


@app.route("/api/collect", methods=["POST"])
def collect_all():
    conn = get_conn()
    printers = [dict(row) for row in conn.execute("SELECT * FROM printers").fetchall()]
    conn.close()
    results, total_saved = [], 0
    for printer in printers:
        content, err = fetch_printer_logs(printer)
        if err:
            results.append({"printer": printer["name"], "ip": printer["ip"], "status": "error", "message": err})
            continue
        logs = parse_csv_content(content, printer)
        saved = save_logs(logs)
        total_saved += saved
        results.append({"printer": printer["name"], "ip": printer["ip"],
                         "status": "ok", "found": len(logs), "saved": saved})
    return jsonify({"results": results, "total_saved": total_saved})


@app.route("/api/collect/<int:pid>", methods=["POST"])
def collect_one(pid):
    conn = get_conn()
    printer_row = conn.execute("SELECT * FROM printers WHERE id = ?", (pid,)).fetchone()
    conn.close()
    if not printer_row:
        return jsonify({"error": "Printer topilmadi"}), 404
    printer = dict(printer_row)
    content, err = fetch_printer_logs(printer)
    if err:
        return jsonify({"status": "error", "message": err})
    logs = parse_csv_content(content, printer)
    saved = save_logs(logs)
    return jsonify({"status": "ok", "found": len(logs), "saved": saved, "printer": printer["name"]})


@app.route("/api/upload-csv", methods=["POST"])
def upload_csv():
    if "file" not in request.files:
        return jsonify({"error": "Fayl yo'q"}), 400
    f = request.files["file"]
    pid = int(request.form.get("printer_id", 0))
    
    conn = get_conn()
    printer_row = conn.execute("SELECT * FROM printers WHERE id = ?", (pid,)).fetchone()
    if not printer_row:
        printer_row = conn.execute("SELECT * FROM printers LIMIT 1").fetchone()
    conn.close()
    
    if not printer_row:
        return jsonify({"error": "Tizimda printerlar topilmadi. Avval printer qo'shing."}), 400
        
    printer = dict(printer_row)
    content = f.read().decode("utf-8-sig", errors="replace")
    logs = parse_csv_content(content, printer)
    saved = save_logs(logs)
    return jsonify({"status": "ok", "found": len(logs), "saved": saved})


@app.route("/api/stats")
def get_stats():
    period = request.args.get("period", "all")  # today | weekly | monthly | all
    now_dt = datetime.now()

    # Davr chegaralarini Python da hisoblash (datetime.now() = mahalliy vaqt, aniq ishonchli)
    # SQLite ning datetime('now','localtime') ba'zan noto'g'ri ishlaydi — shuning uchun
    # chegaralarni Python tomonida hisoblab, tayyor ISO string sifatida beramiz.
    if period == "today":
        today_str = now_dt.strftime("%Y-%m-%d")
        date_filter = f" AND date(created_at) = '{today_str}'"
    elif period == "weekly":
        week_ago = (now_dt - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
        date_filter = f" AND created_at >= '{week_ago}'"
    elif period == "monthly":
        month_ago = (now_dt - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
        date_filter = f" AND created_at >= '{month_ago}'"
    else:
        date_filter = ""

    conn = get_conn()

    c = conn.cursor()

    total_sheets = c.execute(f"SELECT COALESCE(SUM(sheets),0) FROM print_logs WHERE result='OK' AND hidden=0{date_filter}").fetchone()[0]
    total_jobs   = c.execute(f"SELECT COUNT(*) FROM print_logs WHERE hidden=0{date_filter}").fetchone()[0]
    syslog_count = c.execute("SELECT COUNT(*) FROM syslog_raw").fetchone()[0]

    avg_sheets  = c.execute(f"SELECT COALESCE(ROUND(AVG(sheets), 1), 0) FROM print_logs WHERE result='OK' AND hidden=0{date_filter}").fetchone()[0]
    failed_jobs = c.execute(f"SELECT COUNT(*) FROM print_logs WHERE result != 'OK' AND hidden=0{date_filter}").fetchone()[0]

    top_users = c.execute(f"""
        SELECT user_name, SUM(sheets) as total_sheets, COUNT(*) as jobs
        FROM print_logs WHERE result='OK' AND hidden=0{date_filter}
        GROUP BY user_name ORDER BY total_sheets DESC LIMIT 20
    """).fetchall()

    raw_by_printer = c.execute(f"""
        SELECT printer_name, printer_ip, SUM(sheets) as total_sheets, COUNT(*) as jobs
        FROM print_logs WHERE result='OK' AND hidden=0{date_filter}
        GROUP BY printer_ip ORDER BY total_sheets DESC
    """).fetchall()

    daily = c.execute(f"""
        SELECT substr(start_time,1,10) as day, SUM(sheets) as total
        FROM print_logs WHERE result='OK' AND start_time!='' AND hidden=0{date_filter}
        GROUP BY day ORDER BY day DESC LIMIT 30
    """).fetchall()

    by_type = c.execute(f"""
        SELECT job_type, SUM(sheets) as total_sheets, COUNT(*) as jobs
        FROM print_logs WHERE hidden=0{date_filter}
        GROUP BY job_type
    """).fetchall()

    hourly = c.execute(f"""
        SELECT substr(start_time, 12, 2) as hour, COUNT(*) as jobs
        FROM print_logs WHERE start_time!='' AND length(start_time) >= 13 AND hidden=0{date_filter}
        GROUP BY hour ORDER BY hour ASC
    """).fetchall()

    printers = [dict(row) for row in c.execute("SELECT * FROM printers").fetchall()]
    conn.close()

    printers_map = dict((r["ip"], r["name"]) for r in printers)
    by_printer = []
    for r in raw_by_printer:
        ip = r["printer_ip"]
        name = printers_map.get(ip, r["printer_name"] or f"Printer ({ip})")
        by_printer.append({
            "printer_name": name,
            "printer_ip": ip,
            "total_sheets": r["total_sheets"],
            "jobs": r["jobs"]
        })

    return jsonify({
        "total_sheets": total_sheets,
        "total_jobs":   total_jobs,
        "syslog_count": syslog_count,
        "avg_sheets":   avg_sheets,
        "failed_jobs":  failed_jobs,
        "top_users":    [dict(r) for r in top_users],
        "by_printer":   by_printer,
        "daily":        [dict(r) for r in daily],
        "by_type":      [dict(r) for r in by_type],
        "hourly":       [dict(r) for r in hourly],
        "printers":     printers,
        "syslog_port":  SYSLOG_PORT,
        "period":       period,
    })


# ─── PRINTER CRUD API ────────────────────────────────────────────────
@app.route("/api/printers", methods=["GET"])
def get_printers_api():
    conn = get_conn()
    printers = [dict(r) for r in conn.execute("SELECT * FROM printers").fetchall()]
    conn.close()
    return jsonify(printers)

CACHE_FILE = "toner_cache.json"
TONER_CACHE = {}

def load_cache():
    global TONER_CACHE
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r") as f:
                TONER_CACHE = json.load(f)
        except Exception as e:
            log.warning(f"Error loading toner cache file: {e}")
            TONER_CACHE = {}
    else:
        TONER_CACHE = {}

def save_cache():
    try:
        with open(CACHE_FILE, "w") as f:
            json.dump(TONER_CACHE, f, indent=4)
    except Exception as e:
        log.warning(f"Error saving toner cache file: {e}")

# Load cache on startup
load_cache()


def fetch_single_toner(ip):
    session = requests.Session()
    try:
        r_get = session.get(f"http://{ip}:8000/", timeout=5)
        if r_get.status_code != 200:
            return None, f"HTTP status {r_get.status_code}"
    except Exception as e:
        return None, str(e)

    challenge_match = re.search(r'id\s*=\s*"CHALLENGE"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
    if not challenge_match:
        challenge_match = re.search(r'name\s*=\s*"CHALLENGE"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
    challenge = challenge_match.group(1) if challenge_match else None

    pk_match = re.search(r'id\s*=\s*"PK"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
    if not pk_match:
        pk_match = re.search(r'name\s*=\s*"PK"[^>]*value\s*=\s*"([^"]+)"', r_get.text)
    pk_pem = pk_match.group(1) if pk_match else None

    if challenge and pk_pem:
        try:
            import base64
            from cryptography.hazmat.primitives.asymmetric import padding
            from cryptography.hazmat.primitives import serialization
            
            pk_clean = pk_pem.strip().replace("\r", "").replace("\n", "")
            if "BEGIN PUBLIC KEY" not in pk_clean:
                pk_pem_formatted = "-----BEGIN PUBLIC KEY-----\n" + \
                    "\n".join([pk_clean[i:i+64] for i in range(0, len(pk_clean), 64)]) + \
                    "\n-----END PUBLIC KEY-----"
            else:
                m = re.search(r'-----BEGIN PUBLIC KEY-----(.+?)-----END PUBLIC KEY-----', pk_clean, re.DOTALL)
                if m:
                    inner = m.group(1).replace("\n", "").replace(" ", "")
                    pk_pem_formatted = "-----BEGIN PUBLIC KEY-----\n" + \
                        "\n".join([inner[i:i+64] for i in range(0, len(inner), 64)]) + \
                        "\n-----END PUBLIC KEY-----"
                else:
                    pk_pem_formatted = pk_pem

            public_key = serialization.load_pem_public_key(pk_pem_formatted.encode('utf-8'))
            message = (PRINTER_PASS + challenge).encode('utf-8')
            ciphertext = public_key.encrypt(message, padding.PKCS1v15())
            encrypted_pass = base64.b64encode(ciphertext).decode('utf-8')
            session.post(f"http://{ip}:8000/login", data={
                "USERNAME": PRINTER_USER,
                "PASSWORD": encrypted_pass,
                "PASSWORD_T": PRINTER_PASS,
                "CHALLENGE": challenge,
                "URI": "/",
                "policy": "",
                "DOMAIN": "localhost",
                "admin": "",
                "GUEST": "",
                "invalidCH": ""
            }, timeout=5)
        except Exception as e:
            return None, f"Login RSA error: {e}"
    else:
        try:
            session.post(
                f"http://{ip}:8000/login.cgi",
                data={"UserName": PRINTER_USER, "Password": PRINTER_PASS, "Action": "Login"},
                timeout=5
            )
        except Exception as e:
            return None, f"Login.cgi error: {e}"

    try:
        session.get(f"http://{ip}:8000/rps/nativetop.cgi", timeout=5)
    except:
        pass

    try:
        r_main = session.get(f"http://{ip}:8000/", timeout=5)
        match = re.search(r'var\s+tonerVolInfo\s*=\s*({[^}]+})', r_main.text)
        if match:
            vol_data = json.loads(match.group(1))
            return {
                "cyan": int(vol_data.get("tonerCVol", 0)),
                "magenta": int(vol_data.get("tonerMVol", 0)),
                "yellow": int(vol_data.get("tonerYVol", 0)),
                "black": int(vol_data.get("tonerKVol", 0)),
            }, None
        else:
            return None, "tonerVolInfo topilmadi"
    except Exception as e:
        return None, f"Fetch info error: {e}"

def update_toners_loop():
    while True:
        try:
            conn = get_conn()
            printers = [dict(r) for r in conn.execute("SELECT * FROM printers").fetchall()]
            conn.close()
        except Exception as e:
            log.warning(f"Error fetching printers list for toner update: {e}")
            printers = []
            
        threads = []
        def worker(p):
            ip = p["ip"]
            levels, err = fetch_single_toner(ip)
            if levels:
                TONER_CACHE[ip] = {
                    "cyan": levels["cyan"],
                    "magenta": levels["magenta"],
                    "yellow": levels["yellow"],
                    "black": levels["black"],
                    "status": "online",
                    "updated_at": datetime.now().isoformat()
                }
                save_cache()
            else:
                if ip not in TONER_CACHE:
                    TONER_CACHE[ip] = {
                        "cyan": 100, "magenta": 100, "yellow": 100, "black": 100,
                        "status": "offline",
                        "updated_at": None
                    }
                    save_cache()
                else:
                    TONER_CACHE[ip]["status"] = "offline"

        for p in printers:
            t = threading.Thread(target=worker, args=(p,))
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        time.sleep(120)


@app.route("/api/toner", methods=["GET"])
def get_toner_levels():
    conn = get_conn()
    printers = [dict(r) for r in conn.execute("SELECT * FROM printers").fetchall()]
    conn.close()
    
    results = []
    for p in printers:
        ip = p["ip"]
        if ip in TONER_CACHE:
            levels = {
                "black": TONER_CACHE[ip].get("black", 100),
                "cyan": TONER_CACHE[ip].get("cyan", 100),
                "magenta": TONER_CACHE[ip].get("magenta", 100),
                "yellow": TONER_CACHE[ip].get("yellow", 100)
            }
        else:
            levels = {
                "black": 100,
                "cyan": 100,
                "magenta": 100,
                "yellow": 100
            }
        results.append({
            "printer_id": p["id"],
            "name": p["name"],
            "ip": p["ip"],
            "levels": levels
        })
    return jsonify(results)


@app.route("/api/printers", methods=["POST"])
def add_printer_api():
    data = request.json
    name = data.get("name", "").strip()
    ip = data.get("ip", "").strip()
    if not name or not ip:
        return jsonify({"status": "error", "message": "Nom va IP manzili kiritilishi shart"}), 400
    
    conn = get_conn()
    try:
        conn.execute("INSERT INTO printers (name, ip) VALUES (?, ?)", (name, ip))
        conn.execute("UPDATE print_logs SET printer_name = ? WHERE printer_ip = ?", (name, ip))
        conn.commit()
        return jsonify({"status": "ok", "message": "Printer muvaffaqiyatli qo'shildi"})
    except sqlite3.IntegrityError:
        return jsonify({"status": "error", "message": f"IP manzil ({ip}) band yoki avval qo'shilgan"}), 400
    finally:
        conn.close()


@app.route("/api/printers/<int:pid>", methods=["PUT"])
def update_printer_api(pid):
    data = request.json
    name = data.get("name", "").strip()
    ip = data.get("ip", "").strip()
    if not name or not ip:
        return jsonify({"status": "error", "message": "Nom va IP manzili kiritilishi shart"}), 400
    
    conn = get_conn()
    try:
        conn.execute("UPDATE printers SET name = ?, ip = ? WHERE id = ?", (name, ip, pid))
        conn.execute("UPDATE print_logs SET printer_name = ? WHERE printer_ip = ?", (name, ip))
        conn.commit()
        return jsonify({"status": "ok", "message": "Printer muvaffaqiyatli yangilandi"})
    except sqlite3.IntegrityError:
        return jsonify({"status": "error", "message": f"IP manzil ({ip}) boshqa printerga tegishli"}), 400
    finally:
        conn.close()


@app.route("/api/printers/<int:pid>", methods=["DELETE"])
def delete_printer_api(pid):
    conn = get_conn()
    conn.execute("DELETE FROM printers WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "message": "Printer muvaffaqiyatli o'chirildi"})


# Demo API endpoints removed.



@app.route("/api/logs")
def get_logs():
    user       = request.args.get("user", "")
    printer_ip = request.args.get("printer", "")
    source     = request.args.get("source", "")
    limit      = int(request.args.get("limit", 200))
    offset     = int(request.args.get("offset", 0))

    conn = get_conn()
    query  = "SELECT * FROM print_logs WHERE hidden=0"
    params = []
    if user:
        query += " AND user_name LIKE ?"; params.append(f"%{user}%")
    if printer_ip:
        query += " AND printer_ip=?"; params.append(printer_ip)
    if source:
        query += " AND source=?"; params.append(source)
    query += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params += [limit, offset]

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/syslog-feed")
def syslog_feed():
    """Oxirgi 50 ta Syslog raw paketi"""
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM syslog_raw ORDER BY id DESC LIMIT 50"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/clear", methods=["POST"])
def clear_logs():
    conn = get_conn()
    conn.execute("UPDATE print_logs SET hidden = 1")
    conn.execute("DELETE FROM syslog_raw")
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@app.route("/api/export-kpi")
def export_kpi():
    month_str = request.args.get("month")
    if not month_str:
        return jsonify({"error": "month parametri kiritilishi shart (YYYY-MM)"}), 400
        
    try:
        year, month = map(int, month_str.split("-"))
    except:
        return jsonify({"error": "Noto'g'ri oy formati. YYYY-MM ko'rinishida kiriting."}), 400
        
    try:
        from kpi_export import generate_kpi_excel
        out, month_name = generate_kpi_excel(year, month)
        filename = f"KPI_otchet_{month_name}_{year}.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/export-excel")
def export_excel():
    period = request.args.get("period", "all")
    now_dt = datetime.now()

    if period == "monthly":
        try:
            from kpi_export import generate_kpi_excel
            out, month_name = generate_kpi_excel(now_dt.year, now_dt.month)
            filename = f"KPI_otchet_{month_name}_{now_dt.year}.xlsx"
            return send_file(
                out,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            log.warning(f"Monthly KPI export error: {e}")
    
    def parse_date(date_str):
        if not date_str:
            return None
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m %Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue
        # Format "01/06 2026 11:58:11" kabi holatlar uchun re fallback
        m = re.match(r"(\d{2})/(\d{2})\s+(\d{4})\s+(\d{2}):(\d{2}):(\d{2})", date_str.strip())
        if m:
            try:
                return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), int(m.group(4)), int(m.group(5)), int(m.group(6)))
            except ValueError:
                pass
        return None

    conn = get_conn()
    all_logs = conn.execute("SELECT * FROM print_logs WHERE hidden=0 ORDER BY id DESC").fetchall()
    conn.close()

    now_dt = datetime.now()
    filtered_logs = []
    
    for row in all_logs:
        log_dt = parse_date(row["start_time"])
        if not log_dt:
            if period == "all":
                filtered_logs.append(row)
            continue
            
        if period == "daily":
            if log_dt.date() == now_dt.date():
                filtered_logs.append(row)
        elif period == "weekly":
            if (now_dt.date() - log_dt.date()).days <= 7:
                filtered_logs.append(row)
        elif period == "monthly":
            if log_dt.month == now_dt.month and log_dt.year == now_dt.year:
                filtered_logs.append(row)
        else:
            filtered_logs.append(row)

    total_sheets = sum(row["sheets"] for row in filtered_logs if row["result"] == "OK")
    total_jobs = len(filtered_logs)

    user_stats = {}
    for row in filtered_logs:
        if row["result"] == "OK":
            user = row["user_name"]
            user_stats[user] = user_stats.get(user, {"sheets": 0, "jobs": 0})
            user_stats[user]["sheets"] += row["sheets"]
            user_stats[user]["jobs"] += 1

    sorted_users = sorted(user_stats.items(), key=lambda x: x[1]["sheets"], reverse=True)

    # Excel Workbook yaratish
    wb = Workbook()
    
    # 1. Umumiy statistika sahifasi
    ws1 = wb.active
    ws1.title = "Umumiy statistika"
    ws1.views.sheetView[0].showGridLines = True
    
    # Styling elementlari
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    section_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    normal_font = Font(name=font_family, size=11)
    label_font = Font(name=font_family, size=11, bold=True, color="595959")
    
    fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Sarlavha qismi
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "Canon Print Monitor - Hisobot"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center")
    
    # Metadata
    metadata = [
        ("Hisobot turi", period.upper()),
        ("Yaratilgan sana", now_dt.strftime("%d/%m/%Y %H:%M:%S")),
        ("Jami varaqlar", total_sheets),
        ("Jami chop etishlar soni", total_jobs)
    ]
    
    for i, (k, v) in enumerate(metadata, start=3):
        ws1.cell(row=i, column=1, value=k).font = label_font
        ws1.cell(row=i, column=2, value=v).font = normal_font
        if isinstance(v, int):
            ws1.cell(row=i, column=2).number_format = "#,##0"
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal="left")
            
    # Bo'sh joy va Bo'lim sarlavhasi
    ws1.merge_cells("A8:C8")
    ws1["A8"] = "FOYDALANUVCHILAR STATISTIKASI"
    ws1["A8"].font = section_font
    ws1["A8"].fill = fill_title
    ws1["A8"].alignment = Alignment(horizontal="center")
    
    # Jadval sarlavhalari
    headers1 = ["Foydalanuvchi", "Jami varaqlar", "Ishlar soni"]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=9, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
        cell.border = border_thin
        
    # Ma'lumotlarni yozish
    row_idx = 10
    for user, stat in sorted_users:
        c1 = ws1.cell(row=row_idx, column=1, value=user)
        c2 = ws1.cell(row=row_idx, column=2, value=stat["sheets"])
        c3 = ws1.cell(row=row_idx, column=3, value=stat["jobs"])
        
        c1.font = normal_font
        c2.font = normal_font
        c3.font = normal_font
        
        c2.number_format = "#,##0"
        c3.number_format = "#,##0"
        
        c2.alignment = Alignment(horizontal="right")
        c3.alignment = Alignment(horizontal="right")
        
        for c in (c1, c2, c3):
            c.border = border_thin
            if row_idx % 2 == 1:
                c.fill = fill_zebra
        row_idx += 1
        
    # Column auto-fit
    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(9, ws1.max_row + 1):
            val = ws1.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # 2. Batafsil loglar sahifasi
    ws2 = wb.create_sheet(title="Batafsil print loglari")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["No", "Sana/Vaqt", "Printer IP", "Foydalanuvchi", "Fayl nomi", "Varaqlar", "Nusxalar", "Natija", "Manba"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_title
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin
        
    row_idx = 2
    for r in filtered_logs:
        vals = [
            r["job_no"] or "—",
            r["start_time"] or "—",
            r["printer_ip"] or "—",
            r["user_name"] or "Noma'lum",
            r["file_name"] or "—",
            r["sheets"] or 0,
            r["copies"] or 1,
            r["result"] or "OK",
            r["source"] or "syslog"
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.font = normal_font
            cell.border = border_thin
            
            # Alignments
            if col_idx in (1, 2, 3, 8, 9):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (4, 5):
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in (6, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
                
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
        row_idx += 1
        
    # Autofilter o'rnatish
    last_row = len(filtered_logs) + 1
    if last_row > 1:
        ws2.auto_filter.ref = f"A1:I{last_row}"
        
    # Column auto-fit
    for col_idx in range(1, 10):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws2.max_row + 1):
            val = ws2.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Faylni yuborish
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"canon_report_{period}_{now_dt:%Y%m%d}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )



def auto_collect_loop():
    """Har 10 daqiqada barcha printerlardan avtomatik ravishda loglarni yig'ish daemon threadi"""
    log.info("⏰ Avtomatik sinxronlash (har 10 minutda) fon threadi ishga tushdi")
    time.sleep(15)
    while True:
        try:
            conn = get_conn()
            printers = [dict(r) for r in conn.execute("SELECT * FROM printers").fetchall()]
            conn.close()
            
            total_saved = 0
            for printer in printers:
                try:
                    content, err = fetch_printer_logs(printer)
                    if not err and content:
                        logs = parse_audit_csv(content, printer)
                        if not logs:
                            logs = parse_csv_content(content, printer)
                        if logs:
                            saved = save_logs(logs)
                            total_saved += saved
                except Exception as e:
                    log.warning(f"Avto-sinxronlash xatosi ({printer.get('ip')}): {e}")
            if total_saved > 0:
                log.info(f"🔄 Avto-sinxronlash yakunlandi: {total_saved} ta yangi log saqlandi")
            else:
                log.info("🔄 Avto-sinxronlash: yangi loglar topilmadi")
        except Exception as e:
            log.error(f"Avto-sinxronlash loop xatolik: {e}")
        
        # Har 10 daqiqada (600 soniya) takrorlanadi
        time.sleep(600)


# ─── MAIN ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()

    # Syslog serverni fon threadida ishga tushirish
    t = threading.Thread(target=syslog_server, daemon=True)
    t.start()

    # Toner monitoring loopni fon threadida ishga tushirish
    t_toner = threading.Thread(target=update_toners_loop, daemon=True)
    t_toner.start()

    # Avtomatik log sinxronlash (har 10 minutda) fon threadida ishga tushirish
    t_sync = threading.Thread(target=auto_collect_loop, daemon=True)
    t_sync.start()

    print("=" * 55)
    print("  Canon Print Log Monitor ishga tushdi!")
    print(f"  Dashboard:     http://localhost:5000")
    print(f"  Syslog port:   UDP {SYSLOG_PORT}")
    print("  Auto-Sync:     Har 10 daqiqada avtomatik sinxronlash")
    print("=" * 55)
    print()
    print("  Canon printerda sozlash:")
    print("  Remote UI -> Export/Clear Audit Log -> Syslog Settings")
    print("  [OK] Use Syslog Send")
    print(f"  Syslog Server Address: <bu PC ning IP si>")
    print(f"  Syslog Server Port Number: {SYSLOG_PORT}")
    print("  Connection Type: UDP")
    print("=" * 55)

    app.run(debug=False, host="0.0.0.0", port=5000, use_reloader=False)
