import sqlite3
import re
from datetime import datetime
import openpyxl
import io
import os

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "canon_logs.db")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "kpi_template.xlsx")

USERNAME_MAP = {
    "o.atayeva": "Atayeva Ogulnur",
    "a.axadov": "Axadov Asror",
    "a.axmedov": "Axmedov Azizbek",
    "x.axundjanov": "Axundjanov Xumoyunmirzo",
    "u.azamatov": "Azamatov Uchqun",
    "i.babaxanov": "Babaxanov Ilhom",
    "sh.baxriddinov": "Baxriddinov Shamshodbek",
    "sh.baxtiyarov": "Baxtiyarov Shaxzod",
    "b.muhammadsaidov": "Baxtiyor Muxammadsaidov",
    "sh.beknazarov": "Beknazarov Sherzod",
    "s.bobokulov": "Bobokulov Sodiq",
    "d.bulanov": "Bulanov Dilshodbek",
    "m.butayev": "Butayev Ma'ruf",
    "a.abdullayev": "Abdullayev Abbas",
    "a.abduraximov": "Abduraximov Abdulaziz",
    "sh.abdurahmanov": "Abduraxmanov Shuxrat",
    "sh.asatullayeva": "Asatullayeva Shirin",
    "s.solijonov": "Solijonov Sarvar",
    "b.meliyev": "Meliyev Baxtiyor",
    "f.ergashev": "Ergashev Farrux",
    "m.ergashev": "Ergashev Mirzobek",
    "r.eshniyozov": "Eshniyozov Rustambek",
    "i.gaporov": "Gaporov Isfandiyorbek",
    "o.ibroximov": "Ibroximov Otabek",
    "k.isakova": "Isakova Kamila",
    "b.isoyev": "Isoyev Baxodir",
    "i.isomiddinov": "Isomiddinov Ibodullo",
    "f.istamov": "Istamov Firdavs",
    "r.jumaboyev": "Jumaboyev Rashid",
    "m.karimov": "Karimov Ma'murbek",
    "u.karimov": "Karimov Ulug'bek",
    "m.kodirov": "Kodirov Maruf",
    "b.madraximov": "Madraximov Baxtiyorjon",
    "a.mavlonov": "Mavlonov Akmal",
    "o.mirzayev": "Mirzayev Otabek",
    "b.musayev": "Boburjon Musayev",
    "a.muxammatov": "Muxammatov Alisher",
    "o.nabiev": "Nabiev Omonulla",
}

def normalize_name(name):
    if not name: return ""
    return re.sub(r'[^a-zA-Z]', '', str(name)).lower()

def match_user(fish, user_stats):
    if not fish: return None
    fish_norm = normalize_name(fish)
    
    # 1. Exact or mapped username check
    for db_user in user_stats.keys():
        db_user_clean = str(db_user).strip().lower()
        # Direct match with mapped username
        if db_user_clean in USERNAME_MAP:
            mapped = USERNAME_MAP[db_user_clean]
            if normalize_name(mapped) == fish_norm or mapped.lower() in fish.lower() or fish.lower() in mapped.lower():
                return db_user
        
        db_norm = normalize_name(db_user)
        if len(db_norm) > 4 and (db_norm in fish_norm or fish_norm in db_norm):
            return db_user

    # 2. Token / word boundary matching
    parts = str(fish).split()
    if len(parts) >= 2:
        short_fish = normalize_name(parts[0] + parts[1])
        for db_user in user_stats.keys():
            db_norm = normalize_name(db_user)
            if len(db_norm) > 4 and (short_fish in db_norm or db_norm in short_fish):
                return db_user
    return None

def generate_kpi_excel(year, month):
    db_file = DB_PATH if os.path.exists(DB_PATH) else "canon_logs.db"
    conn = sqlite3.connect(db_file)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    logs = c.execute("SELECT * FROM print_logs WHERE result='OK' AND hidden=0").fetchall()
    
    user_stats = {}
    for r in logs:
        date_str = r['start_time']
        if not date_str:
            continue
        log_dt = None
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m %Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                log_dt = datetime.strptime(date_str.strip(), fmt)
                break
            except ValueError:
                pass
        if not log_dt:
            m = re.match(r"(\d{2})/(\d{2})\s+(\d{4})\s+(\d{2}):(\d{2}):(\d{2})", str(date_str).strip())
            if m:
                try:
                    log_dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), int(m.group(4)), int(m.group(5)), int(m.group(6)))
                except ValueError:
                    pass
        
        if log_dt and log_dt.year == year and log_dt.month == month:
            user = r['user_name']
            if user not in user_stats:
                user_stats[user] = {"jobs": 0, "pages": 0}
            user_stats[user]["jobs"] += 1
            user_stats[user]["pages"] += (r['sheets'] or 0)
            
    conn.close()
    
    tmpl_path = TEMPLATE_PATH
    if not os.path.exists(tmpl_path):
        alt_paths = [
            os.path.join(os.getcwd(), "kpi_template.xlsx"),
            r"C:\Users\user\Documents\KPI\KPI_otchet_iyul_pechat_raznye_pokazateli.xlsx",
            r"C:\Users\user\Downloads\KPI_otchet_iyul_pechat_raznye_pokazateli.xlsx"
        ]
        for alt in alt_paths:
            if os.path.exists(alt):
                tmpl_path = alt
                break

    if not os.path.exists(tmpl_path):
        raise Exception(f"Template topilmadi: {TEMPLATE_PATH}")
        
    wb = openpyxl.load_workbook(tmpl_path)
    
    month_names_uz = ["yanvar", "fevral", "mart", "aprel", "may", "iyun", "iyul", "avgust", "sentabr", "oktabr", "noyabr", "dekabr"]
    month_names_ru = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    
    month_name_uz = month_names_uz[month - 1]
    month_name_ru = month_names_ru[month - 1]
    
    # 1. Sheet 0: KPI otchet
    ws_kpi = wb.worksheets[0]
    sheet_title_name = f"KPI otchet {month_name_uz}"
    try:
        ws_kpi.title = sheet_title_name
    except Exception:
        pass

    # Update header title in A2
    ws_kpi.cell(row=2, column=1).value = (
        f"Oʻzbekiston Respublikasi Vazirlar Mahkamasi huzuridagi Yengil sanoatni rivojlantirish agentligida "
        f"{year}-yil {month_name_uz} oyi uchun chop etilgan hujjatlar va sahifalar bo'yicha KPI hisoboti"
    )

    # Detect employee rows starting at row 6
    emp_rows = []
    for r in range(6, ws_kpi.max_row + 1):
        val = ws_kpi.cell(row=r, column=4).value
        val_str = str(val).strip() if val else ""
        if val_str and val_str.upper() not in ("FISH", "F.I.SH.", "JAMI") and not val_str.startswith("Raqamlashtirish"):
            emp_rows.append(r)

    emp_count = len(emp_rows)
    total_jobs = 0
    total_pages = 0
    total_norma = 0
    over_norma_count = 0
    max_pages = 0

    for row_idx in emp_rows:
        fish = ws_kpi.cell(row=row_idx, column=4).value
        try:
            norma = int(ws_kpi.cell(row=row_idx, column=5).value or 220)
        except Exception:
            norma = 220
        
        total_norma += norma
        matched_user = match_user(fish, user_stats)
        
        jobs = 0
        pages = 0
        if matched_user:
            jobs = user_stats[matched_user]["jobs"]
            pages = user_stats[matched_user]["pages"]
            
        total_jobs += jobs
        total_pages += pages
        if pages > max_pages:
            max_pages = pages
            
        ws_kpi.cell(row=row_idx, column=6).value = jobs
        ws_kpi.cell(row=row_idx, column=7).value = pages
        
        ws_kpi.cell(row=row_idx, column=8).value = f"=G{row_idx}/E{row_idx}"
        ws_kpi.cell(row=row_idx, column=9).value = f"=E{row_idx}-G{row_idx}"
        ws_kpi.cell(row=row_idx, column=10).value = "me'yordan oshdi" if pages > norma else "norma doirasida"
        
        if pages > norma:
            over_norma_count += 1

    # Update summary header info box
    formatted_pages = f"{total_pages:,}".replace(",", " ")
    formatted_limit = f"{total_norma:,}".replace(",", " ")
    remaining_limit = max(0, total_norma - total_pages)
    formatted_remaining = f"{remaining_limit:,}".replace(",", " ")

    ws_kpi.cell(row=3, column=1).value = f"JAMI VARAQ: {formatted_pages}"
    ws_kpi.cell(row=3, column=3).value = f"Xodimlar: {emp_count}"
    ws_kpi.cell(row=4, column=3).value = f"Jami limit: {formatted_limit}"
    ws_kpi.cell(row=5, column=3).value = f"Qoldiq limit: {formatted_remaining}"

    # Summary row at bottom
    if emp_rows:
        sum_row = emp_rows[-1] + 1
        ws_kpi.cell(row=sum_row, column=4).value = 'JAMI'
        ws_kpi.cell(row=sum_row, column=5).value = f'=SUM(E6:E{sum_row-1})'
        ws_kpi.cell(row=sum_row, column=6).value = f'=SUM(F6:F{sum_row-1})'
        ws_kpi.cell(row=sum_row, column=7).value = f'=SUM(G6:G{sum_row-1})'
        ws_kpi.cell(row=sum_row, column=8).value = f'=G{sum_row}/E{sum_row}'
        ws_kpi.cell(row=sum_row, column=9).value = f'=SUM(I6:I{sum_row-1})'
        ws_kpi.cell(row=sum_row, column=10).value = 'Normadan oshmagan' if over_norma_count == 0 else f"{over_norma_count} ta me'yordan oshdi"

    # 2. Sheet 1: Сводка
    if len(wb.worksheets) > 1:
        ws_sum = wb.worksheets[1]
        ws_sum.cell(row=1, column=1).value = f"KPI ОТЧЁТ ПО РАСПЕЧАТАННЫМ СТРАНИЦАМ ЗА {month_name_ru.upper()}"
        ws_sum.cell(row=3, column=2).value = month_name_ru
        ws_sum.cell(row=4, column=2).value = emp_count
        ws_sum.cell(row=5, column=2).value = 220
        ws_sum.cell(row=6, column=2).value = '=B4*B5'
        ws_sum.cell(row=7, column=2).value = total_pages
        ws_sum.cell(row=8, column=2).value = '=B6-B7'
        ws_sum.cell(row=9, column=2).value = '=B7/B4'
        ws_sum.cell(row=10, column=2).value = '=B9/B5'
        
        sum_r_end = emp_rows[-1] if emp_rows else 66
        ref_sheet = sheet_title_name.replace("'", "''")
        ws_sum.cell(row=11, column=2).value = f"=MAX('{ref_sheet}'!G6:G{sum_r_end})"
        ws_sum.cell(row=12, column=2).value = f"=COUNTIF('{ref_sheet}'!G6:G{sum_r_end}, \">\"&B5)"
        ws_sum.cell(row=13, column=2).value = f"=B7/SUM('{ref_sheet}'!F6:F{sum_r_end})"
        
        ws_sum.cell(row=3, column=5).value = total_pages
        ws_sum.cell(row=4, column=5).value = total_norma
        ws_sum.cell(row=5, column=5).value = '=E4-E3'

    # 3. Sheet 2: Норма и список
    if len(wb.worksheets) > 2:
        ws_norm = wb.worksheets[2]
        ws_norm.cell(row=4, column=2).value = month_name_ru
        ws_norm.cell(row=5, column=2).value = 220
        ws_norm.cell(row=6, column=2).value = total_pages
        ws_norm.cell(row=7, column=2).value = emp_count

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, month_name_uz

